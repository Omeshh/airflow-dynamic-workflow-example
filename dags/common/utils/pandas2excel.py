import sys
import pandas as pd
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment


class Writer(object):
    def __init__(self, out_file):
        self.out_file = out_file
        self.writer = pd.ExcelWriter(out_file, engine='openpyxl')

    # format string output
    def __str__(self):
        str = ' Pandas to Excel Writer, out_file: ' + self.out_file
        return str

    # create sheet
    def create_sheet(self, sheet_name):
        print('Creating sheet:', sheet_name)
        worksheet = self.writer.book.create_sheet(sheet_name)
        self.writer.sheets[sheet_name] = worksheet

    # write pandas to sheet, row and column are 0-indexed
    def insert_df(self, df, sheet_name, start_row=0, start_col=0, index=True, auto_format=True, header=True, header_axis='both', is_full_border=False):
        if sheet_name not in self.writer.book.sheetnames:
            print('sheet', sheet_name, 'does not exist, creating')
            self.create_sheet(sheet_name)
        df.to_excel(self.writer, sheet_name, startrow=start_row, startcol=start_col, index=index, header=header)
        worksheet = self.writer.sheets[sheet_name]
        if auto_format:
            header_index = str(start_row+1)
            table_end = str(start_row + df.shape[0] + (1 if header else 0))
            start_index = start_col+1
            if index:
                start_index += 1
            end_index = start_index+df.shape[1]-1
            header_font = Font(color='FFFFFF', bold=True)
            header_fill = PatternFill('solid', fgColor='444444')
            thin = Side(border_style="thin", color="000000")
            border = Border(top=thin, left=thin, right=thin, bottom=thin)

            # Set border around table
            to_apply = col_letter(start_index)+header_index+':'+col_letter(end_index)+table_end
            style_range(worksheet, to_apply, border=border, is_full_border=is_full_border)

            if header:
                if header_axis not in ['both', 1, 2]:
                    print('Invalid axis, specify 1 for row, 2 for column')
                    sys.exit()

                if header_axis in ['both', 1]:
                    # Set header style
                    to_apply = col_letter(start_index)+header_index+':'+col_letter(end_index)+header_index
                    style_range(worksheet, to_apply, font=header_font, fill=header_fill)

                    # right align all but first column
                    to_apply = col_letter(start_index)+header_index+':'+col_letter(end_index)+table_end
                    if not index:
                        to_apply = col_letter(start_index+1)+to_apply[1:]
                    style_range(worksheet, to_apply, alignment=Alignment(horizontal='right')) # right align all but first header

                if header_axis in ['both', 2]:
                    # Set first column to match header
                    if index:
                        to_apply = col_letter(start_index-1)+header_index+':'+col_letter(start_index-1)+table_end
                    else:
                        to_apply = col_letter(start_index)+header_index+':'+col_letter(start_index)+table_end
                    style_range(worksheet, to_apply, font=header_font, fill=header_fill, alignment=Alignment(horizontal='left'))

    # format values, axis 1 is row and axis 2 is column
    def format_values(self, sheet_name, axis, format_dictionary, search_index=None):
        workbook = self.writer.book
        worksheet = self.writer.sheets[sheet_name]
        if axis == 1:
            if not search_index:
                search_index = 'A'
            for f in format_dictionary.keys():
                row_index = None
                for i, cell in enumerate(worksheet[search_index+':'+search_index]):
                    if cell.value == f:
                        row_index = cell.row
                        for cell in list(worksheet.rows)[row_index-1]:
                            cell.number_format = number_format(format_dictionary[f])
        elif axis == 2:
            if not search_index:
                search_index = 1
            search_index -= 1 # 0 indexed
            for f in format_dictionary.keys():
                col_index = None
                # find corresponding column given column name
                for i, cell in enumerate(list(worksheet.rows)[search_index]):
                    if cell.value == f:
                        col_index = cell.column
                        for cell in worksheet[col_index+':'+col_index]:
                            cell.number_format = number_format(format_dictionary[f])
        else:
            print('Invalid axis, specify 1 for row and 2 for column')
            sys.exit()

    # paint table, axis 1 is row and axis 2 is column
    def paint_table(self, sheet_name, axis, style_dictionary, search_index=None):
        workbook = self.writer.book
        worksheet = self.writer.sheets[sheet_name]
        if axis == 1:
            if not search_index:
                search_index = 'A'
            for f in style_dictionary.keys():
                row_index = None
                for i, cell in enumerate(worksheet[search_index+':'+search_index]):
                    if cell.value == f:
                        row_index = cell.row
                if row_index:
                    for cell in list(worksheet.rows)[row_index-1]:
                        if cell.value == None:
                            continue
                        cell.fill = fill_styles(style_dictionary[f])
                        cell.font = Font(color='000000', bold=cell.font.bold)
                        if cell.font.bold:
                            if '2' in style_dictionary[f]:
                                cell.alignment = Alignment(indent=2)
                            if '3' in style_dictionary[f]:
                                cell.alignment = Alignment(indent=4)
        elif axis == 2:
            if not search_index:
                search_index = 1
            search_index -= 1 # 0 indexed
            for f in style_dictionary.keys():
                col_index = None
                # find correspond column
                for i, cell in enumerate(list(worksheet.rows)[search_index]):
                    if cell.value == f:
                        col_index = cell.column
                if col_index:
                    for cell in worksheet[col_index+':'+col_index]:
                        if cell.value:
                            cell.fill = fill_styles(style_dictionary[f])
                            cell.font = Font(color='000000', bold=cell.font.bold)
        else:
            print('Invalid axis, specify 1 for row and 2 for column')
            sys.exit()

    # resize all columns automatically based on string length
    def resize_columns(self, sheet_name):
        workbook = self.writer.book
        worksheet = self.writer.sheets[sheet_name]
        for i, col in enumerate(worksheet.columns):
            longest = 0
            longest_type = None
            for cell in col:
                value_type = type(cell.value)
                if value_type == str:
                    string_length = len(str(cell.value))
                elif value_type in [int, float]:
                    string_length = len(str(round(cell.value, 2)))+1
                elif value_type == pd.Timestamp:
                    string_length = 10
                else:
                    string_length = 8
                if string_length > longest:
                    longest = string_length
                    longest_type = value_type
            if longest_type in [int, float]:
                longest = int(1.5*longest)
            elif longest_type == str:
                if longest <= 15:
                    longest = int(1.25*longest)
                elif longest > 25:
                    longest = 25
            worksheet.column_dimensions[col_letter(i+1)].width = longest

    # resize columns from all sheets
    def resize_all(self):
        for sheet in self.writer.sheets:
            self.resize_columns(sheet)

    # remove gridlines
    def remove_gridlines(self, sheet_name):
        workbook = self.writer.book
        worksheet = self.writer.sheets[sheet_name]
        worksheet.sheet_view.showGridLines = False

    # remove gridlines from all sheets
    def remove_all_gridlines(self):
        for sheet in self.writer.sheets:
            self.remove_gridlines(sheet)

    # set widths manually
    def set_widths(self, sheet_name, widths):
        workbook = self.writer.book
        worksheet = self.writer.sheets[sheet_name]
        for w in widths.keys():
            worksheet.column_dimensions[w].width = widths[w]

    def save(self):
        self.writer.save()


def col_letter(i):
    string = ""
    while i > 0:
        i, remainder = divmod(i - 1, 26)
        string = chr(65 + remainder) + string
    return string


def style_range(ws, cell_range, border=Border(), fill=None, font=None, alignment=None, number_format=None, is_full_border=False):
    """
    Apply styles to a range of cells as if they were a single cell.

    :param ws:  Excel worksheet instance
    :param range: An excel range to style (e.g. A1:F20)
    :param border: An openpyxl Border
    :param fill: An openpyxl PatternFill or GradientFill
    :param font: An openpyxl Font object
    """

    top = Border(top=border.top)
    left = Border(left=border.left)
    right = Border(right=border.right)
    bottom = Border(bottom=border.bottom)

    rows = ws[cell_range]

    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        for c in row:
            if is_full_border:
                c.border = border
            if fill:
                c.fill = fill
            if font:
                c.font = font
            if number_format:
                c.number_format = number_format
            if alignment:
                c.alignment = alignment


def number_format(style):
    formats = {
        'integer': '#,##0',
        'numeric': '#,##0.00',
        'percent': '0.00%',
        'dollar': '$#,##0.00',
        'dollar0': '$#,##0',
        'month': 'mmm yyyy',
        'day': 'm/dd/yyyy',
        'datetime': 'dd/mm/yyyy hh:mm AM/PM'
    }
    return formats[style]


def fill_styles(style):
    styles = {
    'Red': PatternFill('solid', fgColor='a82f2f'),
    'Red2': PatternFill('solid', fgColor='E6B4B4'),
    'Red3': PatternFill('solid', fgColor='F0DCDC'),
    'Orange': PatternFill('solid', fgColor='E1640A'),
    'Orange2': PatternFill('solid', fgColor='FABE8C'),
    'Orange3': PatternFill('solid', fgColor='FAE6DC'),
    'Yellow': PatternFill('solid', fgColor='c9c900'),
    'Yellow2': PatternFill('solid', fgColor='F0F0AA'),
    'Yellow3': PatternFill('solid', fgColor='ffffd6'),
    'Green': PatternFill('solid', fgColor='2ca25f'),
    'Green2': PatternFill('solid', fgColor='99d8c9'),
    'Green3': PatternFill('solid', fgColor='e5f5f9'),
    'Blue': PatternFill('solid', fgColor='2b8cbe'),
    'Blue2': PatternFill('solid', fgColor='a6bddb'),
    'Blue3': PatternFill('solid', fgColor='eff3ff'),
    'Purple': PatternFill('solid', fgColor='756bb1'),
    'Purple2': PatternFill('solid', fgColor='bcbddc'),
    'Purple3': PatternFill('solid', fgColor='efedf5'),
    'DarkBlue': PatternFill('solid', fgColor='203764')
    }
    return styles[style]
